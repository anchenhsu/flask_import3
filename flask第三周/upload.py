from flask import Flask, request, render_template, send_file, redirect, url_for, jsonify
from openpyxl import Workbook
import pandas as pd
from jinja2 import Environment
env = Environment()
env.globals.update(zip=zip)
import os 
import MySQLdb

app = Flask(__name__)
app.config['JSON_AS_ASCII'] = False
app.secret_key = os.urandom(24)

conn = MySQLdb.connect(host='localhost', user='root', password='12345678', db='project')

@app.before_first_request
def setup(): 
    app.jinja_env.globals.update(zip=zip)

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST': 
        file = request.files['file']
        df = pd.read_excel(file, engine='openpyxl')
        cursor = conn.cursor()
        for index, row in df.iterrows():
            bmi = round(row[2]/((row[1]/100)**2),2)
            sql = "INSERT INTO project(id, height, weight, BMI) VALUES (%s, %s, %s, %s)"
            cursor.execute(sql, (row[0], row[1], row[2], bmi))
        conn.commit()
        cursor.close()
        return redirect(url_for('table'))
    return render_template('index.html')

@app.route('/search', methods=["GET",'POST'])
def search():
    search_value = request.get_json('searchValue')
    print(search_value)

    cursor = conn.cursor()
    sql = "SELECT * FROM project where id LIKE %s"
    cursor.execute(sql, (search_value['search'],))
    conn.commit()
    data_select = cursor.fetchall()

    result = {'data_select': data_select}
    print(data_select)

    return jsonify(result)

@app.route('/table', methods=["GET",'POST'])
def table():
    cursor = conn.cursor()
    sql = "SELECT * FROM project"
    cursor.execute(sql)
    conn.commit()
    data = cursor.fetchall()    

    return render_template('table.html', data = data)

@app.route('/export', methods=['GET', 'POST'])
def export():
    wb = Workbook()
    ws = wb.active

    cursor = conn.cursor()
    sql = "SELECT * FROM project"
    cursor.execute(sql)
    conn.commit()
    data = cursor.fetchall()

    ws['A1'] = '編號'
    ws['B1'] = '身高'
    ws['C1'] = '體重'
    ws['D1'] = 'BMI'

    for i in range(1, len(data)+1):
        cell = 'A' + str(i+1)
        ws[cell] = data[i-1][0]

        cell = 'B' + str(i+1) 
        ws[cell] = data[i-1][1]

        cell = 'C' + str(i+1)
        ws[cell] = data[i-1][2]

        cell = 'D' + str(i+1)
        ws[cell] = data[i-1][3]

    from io import BytesIO
    file = BytesIO()
    wb.save(file)
    file.seek(0)

    return send_file(file, as_attachment=True, download_name='data.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/export_checkbox', methods = ['POST','GET'])
def export_checkbox():
    wb = Workbook()
    ws = wb.active

    cursor = conn.cursor()
    sql = "SELECT * FROM project" 
    cursor.execute(sql)
    conn.commit()
    db_data = cursor.fetchall()

    # option_list = []
    # for i in range(len(db_data)+1):
    #     option = request.form.get(f'option{i}')
    #     option_list.append(option)

    data = request.get_json()
    if data is not None:
        checkbox = data['checkbox']

    row_num = 2     
    for i in range(len(db_data)+1):
        if checkbox[i]:
            ws['A1'] = '編號'
            cell = 'A' + str(row_num)
            ws[cell] = db_data[i-1][0]

            ws['B1'] = '身高' 
            cell = 'B' + str(row_num)
            ws[cell] = db_data[i-1][1]

            ws['C1'] = '體重'
            cell = 'C' + str(row_num)
            ws[cell] = db_data[i-1][2]

            ws['D1'] = 'BMI'
            cell = 'D' + str(row_num)
            ws[cell] = db_data[i-1][3]

            row_num += 1

    from io import BytesIO
    file = BytesIO()
    wb.save(file)
    file.seek(0)

    return send_file(file, as_attachment=True, download_name='data.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

if __name__ == '__main__':
    app.run(debug=True) 